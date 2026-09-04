import io
import os
import json
import streamlit as st
from PIL import Image

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as OpenPyxlImage
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties


def _apply_a4_print_setup(ws, orientation="portrait"):
    """다운로드한 엑셀을 그대로 인쇄해도 A4 용지 너비에 자동으로 맞춰지도록 설정한다."""
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(left=0.35, right=0.35, top=0.5, bottom=0.5, header=0.2, footer=0.2)

# PDF
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Image as ReportLabImage, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.pdfmetrics import registerFontFamily

# 한글 폰트는 배포 서버(OS)에 따라 설치 여부가 달라 시스템 폰트에 의존하면
# 서버에서 글자가 깨진다(□□□). 리포지토리에 폰트 파일을 직접 포함해 어떤
# 환경(로컬/Streamlit Cloud 등)에서도 동일하게 렌더링되도록 한다.
_FONT_DIR = os.path.join(os.path.dirname(__file__), "..", "assets", "fonts")
_FONT_REGULAR_PATH = os.path.join(_FONT_DIR, "NanumGothic-Regular.ttf")
_FONT_BOLD_PATH = os.path.join(_FONT_DIR, "NanumGothic-Bold.ttf")

FONT_NAME = "Helvetica"
BOLD_FONT_NAME = "Helvetica-Bold"

try:
    pdfmetrics.registerFont(TTFont("NanumGothic", _FONT_REGULAR_PATH))
    pdfmetrics.registerFont(TTFont("NanumGothic-Bold", _FONT_BOLD_PATH))
    registerFontFamily(
        "NanumGothic",
        normal="NanumGothic", bold="NanumGothic-Bold",
        italic="NanumGothic", boldItalic="NanumGothic-Bold",
    )
    FONT_NAME = "NanumGothic"
    BOLD_FONT_NAME = "NanumGothic-Bold"
except Exception:
    # 번들 폰트 로드가 실패하면(파일 누락 등) 시스템 폰트로 최후 폴백한다.
    for p in [
        "/System/Library/Fonts/Supplemental/AppleGothic.ttf",
        "/Library/Fonts/AppleGothic.ttf",
        "C:/Windows/Fonts/malgun.ttf",
        "C:/Windows/Fonts/gulim.ttc",
        "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
    ]:
        if os.path.exists(p):
            try:
                pdfmetrics.registerFont(TTFont("CustomKorean", p))
                FONT_NAME = "CustomKorean"
                BOLD_FONT_NAME = "CustomKorean"
                break
            except Exception:
                pass

# 체크리스트 기본 항목 정의 (사진 속 원본 항목 100% 일치)
CHECKLIST_SECTIONS = [
    {
        "category": "1. 차량외부",
        "items": [
            ("파손(도장) 및 CI/BI 상태", "각 부위별 파손 및 외부 충격에 의한 도장 벗겨짐이 없어야 한다. 차량의 외관의 도장 벗겨짐(1cm이상), 파임, 뜸, 벗겨짐(0.5cm)", "item_ext_1"),
            ("등화장치 및 방향 지시기", "점멸작용이 확실하고 파손됨이 없는가?", "item_ext_2"),
            ("후사경 및 반사경", "표면이 깨끗하고 비침 상태가 적정한가?", "item_ext_3"),
            ("경음기 및 와이퍼", "크락션이 소리가 작동하는가?", "item_ext_4"),
            ("등록번호표", "손실/손상이 없고 육안으로 식별이 가능한가?", "item_ext_5"),
            ("슬라이딩짐틀", "앞/뒤 조작이 가능한가?", "item_ext_6"),
            ("슬라이딩짐틀", "외부 충격에 파손이 없는가? (파손, 깨짐, 시야가 가리는 기스)", "item_ext_7"),
            ("스크린미들(전면 바람막이)", "외부 충격에 파손이 없는가? (파손, 깨짐, 시야가 가리는 기스)", "item_ext_8"),
        ]
    },
    {
        "category": "2. 조향장치",
        "items": [
            ("핸들", "출발 전 심한 유동이나 흔들림 떨림 현상이 있는가?", "item_str_1"),
            ("핸들", "주행시 이상하게 잡히거나 무겁지 아니한가?", "item_str_2"),
            ("핸들", "저속 주행시 상하로 덜림 현상은 없는가?", "item_str_3"),
            ("핸들", "브레이크 작동 시 핸들이 떨리는 현상은 없는가?", "item_str_4"),
        ]
    },
    {
        "category": "3. 제동장치",
        "items": [
            ("브레이크", "핸들 브레이크를 동작했을 때 간격이 적당하며, 제동작동이 양호한가?", "item_brk_1"),
        ]
    },
    {
        "category": "4. 주행장치",
        "items": [
            ("타이어", "타이어의 공기압이 적당한가?", "item_tire_1"),
            ("타이어", "이상마모 되거나 균열 또는 손상은 없는가?", "item_tire_2"),
        ]
    },
    {
        "category": "5. 기  타",
        "items": [
            ("블랙박스", "카메라 각도는 전면, 측면을 촬영하고 있는가?", "item_etc_1"),
            ("누적km", "점검일자 기준으로 누적 키로 수 확인 및 작성", "item_km"),
            ("보관함", "각 부위별 파손 및 외부 충격에 의한 도장 벗겨짐이 없어야 한다. 차량의 외관의 도장 벗겨짐(1cm이상), 파임, 뜸, 벗겨짐(0.5cm)", "item_etc_2"),
            ("블루투스이어폰", "외부 충격에 의한 파손은 없는가?", "item_etc_3"),
            ("블루투스이어폰", "송/수신에 문제는 없는가?", "item_etc_4"),
            ("전일 이상 했던 부분", "해당 부분이 정상이어야 한다.", "item_etc_5"),
        ]
    }
]

def compress_image_bytes(raw_bytes, max_dim=1600, quality=82):
    """DB에 원본 그대로 저장하지 않고, 긴 변 기준 max_dim으로 축소 후 JPEG로 압축한다."""
    im = Image.open(io.BytesIO(raw_bytes)).convert("RGB")
    im.thumbnail((max_dim, max_dim), Image.Resampling.LANCZOS)
    out = io.BytesIO()
    im.save(out, format="JPEG", quality=quality, optimize=True)
    return out.getvalue()


def _fit_image(raw_bytes, box_w, box_h):
    """원본 종횡비를 유지한 채 box_w x box_h 흰 배경 캔버스 중앙에 맞춰 넣는다."""
    im = Image.open(io.BytesIO(raw_bytes)).convert("RGB")
    im.thumbnail((box_w, box_h), Image.Resampling.LANCZOS)
    canvas = Image.new("RGB", (box_w, box_h), "white")
    canvas.paste(im, ((box_w - im.width) // 2, (box_h - im.height) // 2))
    return canvas


@st.cache_data(show_spinner=False)
def generate_integrated_excel(row_data):
    wb = Workbook()
    
    # [Sheet 1] 이륜차량 안전관리 상태 평가 (체크리스트)
    ws1 = wb.active
    ws1.title = "안전관리_점검표"
    ws1.views.sheetView[0].showGridLines = True
    _apply_a4_print_setup(ws1, orientation="portrait")

    # [Sheet 2] 차량 4면 사진 평가
    ws2 = wb.create_sheet(title="4면_사진평가")
    ws2.views.sheetView[0].showGridLines = True
    _apply_a4_print_setup(ws2, orientation="landscape")
    
    # 얇은 테두리
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # === Sheet 1 채우기 ===
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 24
    ws1.column_dimensions['C'].width = 50
    ws1.column_dimensions['D'].width = 18
    ws1.column_dimensions['E'].width = 24

    # 타이틀
    ws1.merge_cells("A2:D2")
    ws1["A2"] = "이륜차량 안전관리 상태 평가"
    ws1["A2"].font = Font(name="맑은 고딕", size=18, bold=True)
    ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 36

    # 헤더 메타 (본부명/차량번호는 값이 길어 잘리지 않도록 2개 열을 병합해 공간을 확보한다)
    ws1.merge_cells("A4:B4")
    ws1["A4"] = f"● 본부명 : {row_data['hq_name']}"
    ws1["C4"] = f"● 지사명 : {row_data['branch_name']}"
    ws1.merge_cells("D4:E4")
    ws1["D4"] = f"● 차량번호 : {row_data['car_no']}"
    for c in ["A4", "C4", "D4"]:
        ws1[c].font = Font(name="맑은 고딕", size=11, bold=True)
    ws1.row_dimensions[4].height = 24

    # 표 헤더
    ws1["A5"] = "점 검 부 분"
    ws1.merge_cells("B5:C5")
    ws1["B5"] = "점    검    내    용"
    ws1["D5"] = "결 과"
    for col_l in ["A5", "B5", "C5", "D5"]:
        ws1[col_l].font = Font(name="맑은 고딕", size=10, bold=True)
        ws1[col_l].alignment = Alignment(horizontal="center", vertical="center")
        ws1[col_l].fill = header_fill
        ws1[col_l].border = thin_border
    ws1.row_dimensions[5].height = 24

    check_dict = json.loads(row_data.get("check_data", "{}"))
    cur_r = 6
    
    for sec in CHECKLIST_SECTIONS:
        sec_start = cur_r
        cat_name = sec["category"]
        for sub_cat, desc, key in sec["items"]:
            ws1.cell(row=cur_r, column=2, value=sub_cat).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws1.cell(row=cur_r, column=3, value=f"ㅇ {desc}").alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # 결과 표시
            res_val = check_dict.get(key, "적정")
            if key == "item_km":
                res_cell_val = f"km : {row_data.get('accumulated_km', '')}"
            else:
                res_cell_val = "■ 적정   □ 정비필요" if res_val == "적정" else "□ 적정   ■ 정비필요"
            
            ws1.cell(row=cur_r, column=4, value=res_cell_val).alignment = Alignment(horizontal="center", vertical="center")
            
            for c in range(1, 5):
                ws1.cell(row=cur_r, column=c).border = thin_border
                ws1.cell(row=cur_r, column=c).font = Font(name="맑은 고딕", size=10)
            ws1.row_dimensions[cur_r].height = 32
            cur_r += 1
            
        sec_end = cur_r - 1
        ws1.merge_cells(start_row=sec_start, start_column=1, end_row=sec_end, end_column=1)
        ws1.cell(row=sec_start, column=1, value=cat_name).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 하단 점검일시 및 점검자 서명란
    cur_r += 1
    ws1.merge_cells(start_row=cur_r, start_column=3, end_row=cur_r, end_column=4)
    ws1.cell(row=cur_r, column=3, value=f"점검일시 :  {row_data.get('inspect_date', '')}").font = Font(name="맑은 고딕", size=11, bold=True)
    cur_r += 1
    ws1.merge_cells(start_row=cur_r, start_column=3, end_row=cur_r, end_column=4)
    ws1.cell(row=cur_r, column=3, value=f"점 검 자 :  {row_data.get('inspector', '')} (서명)").font = Font(name="맑은 고딕", size=11, bold=True)

    # === Sheet 2 채우기 (사진 4면) ===
    for c in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws2.column_dimensions[c].width = 11
        
    ws2.merge_cells("A2:H2")
    ws2["A2"] = "기술/업무용 차량 안전관리 상태 평가"
    ws2["A2"].font = Font(name="맑은 고딕", size=18, bold=True)
    ws2["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[2].height = 36

    ws2.merge_cells("A4:B4")
    ws2["A4"] = f"● 본부명 : {row_data['hq_name']}"
    ws2.merge_cells("D4:E4")
    ws2["D4"] = f"● 지사명 : {row_data['branch_name']}"
    ws2.merge_cells("G4:H4")
    ws2["G4"] = f"● 차량번호 : {row_data['car_no']}"
    for c in ["A4", "D4", "G4"]:
        ws2[c].font = Font(name="맑은 고딕", size=11, bold=True)
    ws2.row_dimensions[4].height = 24

    def draw_photo_box(cs, rs, ce, re, label):
        ws2.merge_cells(start_row=rs, start_column=cs, end_row=rs, end_column=ce)
        cell = ws2.cell(row=rs, column=cs, value=label)
        cell.font = Font(name="맑은 고딕", size=11, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
        for r in range(rs, re + 1):
            for c in range(cs, ce + 1):
                ws2.cell(row=r, column=c).border = thin_border

    draw_photo_box(1, 6, 4, 21, "전면")
    draw_photo_box(5, 6, 8, 21, "후면")
    draw_photo_box(1, 23, 4, 38, "우측면")
    draw_photo_box(5, 23, 8, 38, "좌측면")

    img_map = {
        "front": ("A7", row_data.get("img_front")),
        "rear": ("E7", row_data.get("img_rear")),
        "right": ("A24", row_data.get("img_right")),
        "left": ("E24", row_data.get("img_left"))
    }
    for pos, raw_b in img_map.values():
        if raw_b:
            im = _fit_image(raw_b, 310, 270)
            tmp_io = io.BytesIO()
            im.save(tmp_io, format="PNG")
            tmp_io.seek(0)
            ws2.add_image(OpenPyxlImage(tmp_io), pos)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

@st.cache_data(show_spinner=False)
def generate_integrated_pdf(row_data):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=25
    )
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(name="T1", fontName=BOLD_FONT_NAME, fontSize=16, alignment=1, spaceAfter=18)
    meta_style = ParagraphStyle(name="M1", fontName=FONT_NAME, fontSize=10, leading=16)
    tbl_font = ParagraphStyle(name="TB", fontName=FONT_NAME, fontSize=8, leading=12.5)
    tbl_font_center = ParagraphStyle(name="TBC", fontName=FONT_NAME, fontSize=8.5, leading=12.5, alignment=1)

    # === [PAGE 1: 이륜차량 안전관리 상태 평가] ===
    elements.append(Paragraph("이륜차량 안전관리 상태 평가", title_style))

    meta_p1 = Table([[
        Paragraph(f"● 본부명 : {row_data['hq_name']}", meta_style),
        Paragraph(f"● 지사명 : {row_data['branch_name']}", meta_style),
        Paragraph(f"● 차량번호 : {row_data['car_no']}", meta_style)
    ]], colWidths=[180, 160, 195])
    meta_p1.setStyle(TableStyle([('BOTTOMPADDING', (0,0), (-1,-1), 8)]))
    elements.append(meta_p1)
    elements.append(Spacer(1, 8))

    table_data = [
        [
            Paragraph("<b>점검부분</b>", tbl_font_center),
            Paragraph("<b>점  검  내  용</b>", tbl_font_center),
            "",
            Paragraph("<b>결 과</b>", tbl_font_center)
        ]
    ]

    check_dict = json.loads(row_data.get("check_data", "{}"))
    spans = [('SPAN', (1, 0), (2, 0))]
    cur_idx = 1

    for sec in CHECKLIST_SECTIONS:
        start_row = cur_idx
        for sub_cat, desc, key in sec["items"]:
            res_val = check_dict.get(key, "적정")
            if key == "item_km":
                res_txt = f"km : {row_data.get('accumulated_km', '')}"
            else:
                res_txt = "■ 적정  □ 정비필요" if res_val == "적정" else "□ 적정  ■ 정비필요"

            table_data.append([
                Paragraph(f"<b>{sec['category']}</b>", tbl_font_center),
                Paragraph(sub_cat, tbl_font_center),
                Paragraph(f"ㅇ {desc}", tbl_font),
                Paragraph(res_txt, tbl_font_center)
            ])
            cur_idx += 1
        end_row = cur_idx - 1
        spans.append(('SPAN', (0, start_row), (0, end_row)))

    p1_table = Table(table_data, colWidths=[65, 95, 275, 100])
    p1_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), FONT_NAME),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.7, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.whitesmoke),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING', (0,0), (-1,-1), 5),
        ('RIGHTPADDING', (0,0), (-1,-1), 5),
    ] + spans))
    elements.append(p1_table)
    elements.append(Spacer(1, 12))

    sign_data = [
        ["", Paragraph(f"점검일시 : &nbsp;&nbsp;&nbsp;&nbsp; {row_data.get('inspect_date', '')}", meta_style)],
        ["", Paragraph(f"점검자 : &nbsp;&nbsp;&nbsp;&nbsp; {row_data.get('inspector', '')} (서명)", meta_style)]
    ]
    sign_table = Table(sign_data, colWidths=[330, 205])
    sign_table.setStyle(TableStyle([
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ('TOPPADDING', (0,0), (-1,-1), 2),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2)
    ]))
    elements.append(sign_table)

    # === [PAGE 2: 기술/업무용 차량 안전관리 상태 평가 (4면 사진)] ===
    elements.append(PageBreak())
    elements.append(Paragraph("기술/업무용 차량 안전관리 상태 평가", title_style))
    elements.append(meta_p1)
    elements.append(Spacer(1, 10))

    def make_rl_img(b):
        if b:
            fitted = _fit_image(b, 260, 240)
            tmp_io = io.BytesIO()
            fitted.save(tmp_io, format="PNG")
            tmp_io.seek(0)
            return ReportLabImage(tmp_io, width=260, height=240)
        return Paragraph("사진 미등록", meta_style)

    p2_grid = [
        ["전면", "후면"],
        [make_rl_img(row_data.get("img_front")), make_rl_img(row_data.get("img_rear"))],
        ["우측면", "좌측면"],
        [make_rl_img(row_data.get("img_right")), make_rl_img(row_data.get("img_left"))]
    ]
    p2_table = Table(p2_grid, colWidths=[267, 267], rowHeights=[24, 250, 24, 250])
    p2_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), FONT_NAME),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.8, colors.black),
        ('BACKGROUND', (0,0), (1,0), colors.whitesmoke),
        ('BACKGROUND', (0,2), (1,2), colors.whitesmoke),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING', (0,0), (-1,-1), 0),
        ('FONTSIZE', (0,0), (-1,-1), 11),
        ('TOPPADDING', (0,0), (1,0), 6),
        ('BOTTOMPADDING', (0,0), (1,0), 6),
        ('TOPPADDING', (0,2), (1,2), 6),
        ('BOTTOMPADDING', (0,2), (1,2), 6),
    ]))
    elements.append(p2_table)

    doc.build(elements)
    buffer.seek(0)
    return buffer
